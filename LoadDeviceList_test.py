from typing import List, Any, Tuple
from openpyxl import load_workbook
"""
从EXCEL文件中读取IP/IDRAC_IP, 主机名、用户名、第一层密码、第二层密码、品牌
filename1: 定义EXCEL格式的文件，获取数据并返回一个list
sheetname1: 工作表(sheet)中保存交换机、密码等信息的sheet名称
vlanIP_Row1： VLAN_IP: 第3列/C
idracipRow1： IDRAC_IP: 第4列/D
hostname: 主机名：第5列/E
username: 用户名：第6列/F
passwd1Row: 第一层密码：第7列/G
password2Row: 第二层密码：第8列/H
devicebrand:品牌类型：第13列/M ，目前取值仅限CISCO和H3C
程序实现LoadDeviceList.py
"""

_DeviceIP: List[Any] = []
def LoadData(_filename, _sheetname):
    _idrac_row = "D"
    ipadr = ""
    _vlanIP_Row1 = int(3)
    _hostname_row = int(5)
    _username_row = int(6)
    _passwd1Row = int(7)
    _password2Row = int(8)
    _devicebrand = int(13)
    _wb = load_workbook(_filename)
    _sheet: object = _wb.get_sheet_by_name(_sheetname)
    hn=""
    un=""
    passwd1=""
    passwd2=""
    devicebrand=""
    for IP in _sheet[_idrac_row]:
        if IP.row != 1:  # 跳过标题行
            if not None == IP.value:  # 有登记IDRAC ip地址，则以IDRAC地址为准，否则取登记的VLAN IP地址
                ipadr = IP.value
                hn = _sheet.cell(IP.row, _hostname_row).value    #Get device name
                un = _sheet.cell(IP.row, _username_row).value
                passwd1 = _sheet.cell(IP.row, _passwd1Row).value   #第1个密码在第5列
                passwd2 = _sheet.cell(IP.row, _password2Row).value    #第2个密码在第6列
                devicebrand = _sheet.cell(IP.row, _devicebrand).value
            else:
                # IDRAC IP 为空，取前一列的IP地址。
                ipadr = _sheet.cell(IP.row, _vlanIP_Row1).value  #Get IP ADDRESS of vlan
                hn = _sheet.cell(IP.row, _hostname_row).value
                un = _sheet.cell(IP.row, _username_row).value
                passwd1 = _sheet.cell(IP.row, _passwd1Row).value   #第1个密码在第5列
                passwd2 = _sheet.cell(IP.row, _password2Row).value    #第2个密码在第6列
                devicebrand = _sheet.cell(IP.row, _devicebrand).value
            if ipadr is not None:
             _DeviceIP.append([ipadr, hn, un, passwd1, passwd2, devicebrand])
    return _DeviceIP
